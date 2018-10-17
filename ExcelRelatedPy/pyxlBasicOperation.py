#! python3

# アクティブシート
import openpyxl
wb = openpyxl.Workbook()
ws = wb.active


# 新規シート作成

import openpyxl
wb = openpyxl.Workbook()
ws = wb.active
ws = wb.create_sheet()                                      # 右端に作成
ws = wb.create_sheet(0)                                     # 左端に作成
ws = wb.create_sheet(title='SHEETNAME')                     # シート名を指定


# シート名から選択
ws = wb.get_sheet_by_name('SHEETNAME')

# ワークシート名一覧（リスト）
sheetnames = wb.get_sheet_names()


# シート名変更
ws.title = 'New Title'


# セル基本操作:セル書き込み

ws['A1'].value = 'Hello World'

from datetime import datetime as dt

ws.cell(row=2, column=1).value = dt.now()


# セル複数書き込み
i = 1234.5678
for row in ws.iter_rows('C1:E3'):
    for cell in row:
        cell.value = i
        i = i * 2


# 行または列の選択
ws.column_dimensions['A']
ws.row_dimensions[1]


# セル設定変更: セル結合

ws.merge_cells('A5:B5')
ws.unmerge_cells('A5:B5')
ws.merge_cells(start_row=5,start_column=1,end_row=5,end_column=6)
ws.unmerge_cells(start_row=5,start_column=1,end_row=5,end_column=6)


# cell コメント
from openpyxl.comments import Comment

ws['C1'].comment = Comment(
    'This is the comment text',
    'Comment Author'
)

ws['C1'].comment.text  # コメント内容
ws['C1'].comment.author  # コメント作成者


# cellフォント

from openpyxl.styles import Font

ws['A1'].font = Font(
    name='Calibri',
    size=12,
    bold=True,
    italic=False,
    color='FF000000'
)


# セル罫線

from openpyxl.styles import Border, Side

#囲み
ws['B10'].border = Border(
    outline=True,
    left=Side(style='thin', color='FF000000'),
    right=Side(style='thick', color='FF000000'),
    top=Side(style='mediam, color='FF000000'),
    bottom=Side(style=None, color='FF000000')
)

# 斜め
ws['B12'].border = Border(
    outline=True,
    diagonalUp=True,
    diagonalDown=True,
    diagonal=Side(style=None, color='FF000000')
)

style = [
    'dashDot', 'dashDotDot', 'dashed', 'dotted',
    'double', 'hair',
    'medium', 'mediumDashDot', 'mediumDashDotDot', 'mediumDashed',
    'slantDashDot',
    'thick', 'thin'
]













