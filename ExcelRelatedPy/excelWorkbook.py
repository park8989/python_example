#! python3
# cellに値（value）を記入してexcelファイルを作成して、いまのフォルダーにファイルが生成されていることを確認する

from openpyxl import Workbook
import glob

wb = Workbook()
ws = wb.active
ws.cell(row = 2, column = 2).value = 777       #row: A1 A2 A3, 行の変化、 column: A1 B1 C1、列の変化, または、ws['B2'] = 777
wb.save('sample.xlsx')                         #sample.xlsxのファイル名で保存
print(glob.glob("*.xlsx"))                     # xlsxの拡張子を持つファイルを表示



# excelのB列のdataをすべて取得する

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
for i in range(1, 8):
  print(sheet.cell(row=i, column=2).value)
  

#上では行数を指定しているが、指定なしに、列や行の値(value)をまとめて取得する

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
for cell_obj in list(sheet.column)[1]:        # [1]はB列を表す
  print(cell_obj.value)
  


# 新規にexcelファイルを作成する。
import openpyxl
import glob

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'test_sheet_1'
wb.save("test.xlsx")
glob.glob("*.xlsx")



# cellへの値(value)の書き込み

import openpyxl
import glob

wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 'test'
wb.save(test_write.xlsx')
glob.glob("*.xlsx")


# excelに数式を入力する

import openpyxl
import glob

wb = openpyxl.Workbook()
sheet = wb.active
sheet[`A1`] = 10
sheet[`A2`] = 20
sheet[`A3`] = 30
sheet[`A4`] = `=SUM(A1:A3)`
wb.save('test_write_4.xlsx`)


# ほかの方法 enumerateを使う：enumerateの使い方 
# forループでインデックスを取得できるenumerate関数
# enumerateを使うと、forループの中でリスト（配列）などのイテラブルオブジェクトの要素と同時にインデックス番号（カウント、順番）を取得できる。
# https://note.nkmk.me/python-enumerate-start/

# 通常のループ
l = ['Alice', 'Bob', 'Charlie']

for name in l:
    print(name)
# Alice
# Bob
# Charlie

# enumerateを使うと

l = ['Alice', 'Bob', 'Charlie']
for i, name in enumerate(l):
    print(i, name)
# 0 Alice
# 1 Bob
# 2 Charlie


# enumerate関数のインデックスを1から開始

l = ['Alice', 'Bob', 'Charlie']
for i, name in enumerate(l, 1):
    print(i, name)
# 1 Alice
# 2 Bob
# 3 Charlie


# インデックスの増分（step）を指定

l = ['Alice', 'Bob', 'Charlie']
step = 3
for i, name in enumerate(l):
    print(i * step, name)
# 0 Alice
# 3 Bob
# 6 Charlie


# excelへenumerate適用

import openpyxl

numbers = [10, 20, 30]
for i, number in enumerate(enumerate(numbers):
  sheet.cell(row=i+1, column=1).value = number
sheet['A4'] = sum(number)



